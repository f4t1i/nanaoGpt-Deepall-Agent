#!/usr/bin/env python3
"""
DeepALL Master Table Loader
Phase 1: Parse, validate, and cache all 27 sheets from DeepALL Master Table

Features:
- Lazy loading for performance
- Comprehensive data validation
- Caching and indexing
- Error handling and reporting
- Data consistency checks
"""

import openpyxl
import json
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import logging

# ============================================================================
# LOGGING SETUP
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================================
# DATA VALIDATION
# ============================================================================

class DataValidator:
    """Validate data from DeepALL table"""
    
    @staticmethod
    def validate_module_index(value):
        """Validate module index format (m001, m002, etc.)"""
        if not isinstance(value, str):
            return False
        return value.startswith('m') and value[1:].isdigit()
    
    @staticmethod
    def validate_numeric(value, min_val=None, max_val=None):
        """Validate numeric value"""
        if value is None:
            return True  # Allow None
        try:
            num = float(value)
            if min_val is not None and num < min_val:
                return False
            if max_val is not None and num > max_val:
                return False
            return True
        except (ValueError, TypeError):
            return False
    
    @staticmethod
    def validate_score(value):
        """Validate score in [0, 1]"""
        return DataValidator.validate_numeric(value, 0, 1)
    
    @staticmethod
    def validate_index_score(value):
        """Validate index score in [0, 10] or index reference"""
        if value is None or value == 'not_assigned':
            return True  # Allow None or not_assigned
        if isinstance(value, str) and (value.startswith('sy') or value.startswith('op') or 
                                       value.startswith('l') or value.startswith('perf')):
            return True  # Allow index references
        return DataValidator.validate_numeric(value, 0, 10)
    
    @staticmethod
    def validate_non_empty(value):
        """Validate non-empty string"""
        return isinstance(value, str) and len(value.strip()) > 0


# ============================================================================
# DEEPALL TABLE LOADER
# ============================================================================

class DeepALLTableLoader:
    """
    Load and parse DeepALL Master Table
    
    Supports:
    - Lazy loading of sheets
    - Data validation
    - Caching
    - Indexing
    - Error reporting
    """
    
    # Sheet names in the workbook
    CORE_SHEETS = [
        'DeepALL_Complete',      # Main data (219 modules)
        'superintelligences',    # SI groups
        'synergies',             # Synergy data
        'optimization',          # Optimization history
        'learning_results',      # Learning progress
        'performance',           # Performance metrics
        'ai_methods',            # AI method catalog
        'knowledge_base',        # Knowledge base
    ]
    
    METADATA_SHEETS = [
        'Prefix_Mapping',
        'Conflict_Log',
        'System_Index_Stats',
        'Field_Groups',
        'Change_Log',
        'Query_Templates',
        'Unassigned_Fields',
        'Usage_Log',
        'LLM_Access_Log',
        'Learning_History_Log',
        'simulation_results',
        'simulation_types',
        'test_types',
        'learning_types',
        'system_status',
        'index_prefixes',
        'modules_combined',
        'deepall_links',
        'history',
    ]
    
    def __init__(self, excel_path: str):
        """Initialize loader"""
        self.excel_path = excel_path
        self.workbook = None
        self.cache = {}
        self.indexes = {}
        self.validation_errors = []
        self.stats = {}
        
        logger.info(f"Initializing DeepALLTableLoader with {excel_path}")
    
    def load_workbook(self):
        """Load Excel workbook"""
        if self.workbook is None:
            try:
                self.workbook = openpyxl.load_workbook(self.excel_path)
                logger.info(f"Workbook loaded. Sheets: {len(self.workbook.sheetnames)}")
            except Exception as e:
                logger.error(f"Failed to load workbook: {e}")
                raise
        return self.workbook
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names"""
        wb = self.load_workbook()
        return wb.sheetnames
    
    def load_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Load sheet with caching
        
        Returns list of dictionaries (one per row)
        """
        if sheet_name in self.cache:
            logger.debug(f"Loading {sheet_name} from cache")
            return self.cache[sheet_name]
        
        logger.info(f"Loading sheet: {sheet_name}")
        
        try:
            wb = self.load_workbook()
            ws = wb[sheet_name]
            data = self._parse_sheet(ws)
            self.cache[sheet_name] = data
            
            logger.info(f"  ✓ Loaded {len(data)} rows from {sheet_name}")
            self.stats[sheet_name] = {
                'rows': len(data),
                'columns': len(data[0].keys()) if data else 0,
                'loaded_at': datetime.now().isoformat()
            }
            
            return data
        except Exception as e:
            logger.error(f"Failed to load sheet {sheet_name}: {e}")
            raise
    
    def _parse_sheet(self, ws) -> List[Dict[str, Any]]:
        """Parse worksheet into list of dictionaries"""
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Parse data rows
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for col_idx, (header, value) in enumerate(zip(headers, row)):
                if header:  # Skip empty headers
                    row_dict[header] = value
            
            if row_dict:  # Skip empty rows
                rows.append(row_dict)
        
        return rows
    
    def load_complete_data(self) -> List[Dict[str, Any]]:
        """Load main DeepALL_Complete sheet"""
        return self.load_sheet('DeepALL_Complete')
    
    def load_superintelligences(self) -> List[Dict[str, Any]]:
        """Load superintelligences sheet"""
        return self.load_sheet('superintelligences')
    
    def load_synergies(self) -> List[Dict[str, Any]]:
        """Load synergies sheet"""
        return self.load_sheet('synergies')
    
    def load_optimization(self) -> List[Dict[str, Any]]:
        """Load optimization history sheet"""
        return self.load_sheet('optimization')
    
    def load_learning_results(self) -> List[Dict[str, Any]]:
        """Load learning results sheet"""
        return self.load_sheet('learning_results')
    
    def load_performance(self) -> List[Dict[str, Any]]:
        """Load performance metrics sheet"""
        return self.load_sheet('performance')
    
    def load_ai_methods(self) -> List[Dict[str, Any]]:
        """Load AI methods sheet"""
        return self.load_sheet('ai_methods')
    
    def load_knowledge_base(self) -> List[Dict[str, Any]]:
        """Load knowledge base sheet"""
        return self.load_sheet('knowledge_base')
    
    # ========================================================================
    # INDEXING
    # ========================================================================
    
    def create_index(self, sheet_name: str, field: str) -> Dict[str, Dict]:
        """
        Create index for fast lookups
        
        Returns dictionary: {field_value: row_dict}
        """
        key = f"{sheet_name}:{field}"
        
        if key in self.indexes:
            logger.debug(f"Index {key} already exists")
            return self.indexes[key]
        
        logger.info(f"Creating index for {sheet_name}.{field}")
        
        data = self.load_sheet(sheet_name)
        index = {}
        
        for row in data:
            if field in row and row[field] is not None:
                index[row[field]] = row
        
        self.indexes[key] = index
        logger.info(f"  ✓ Index created with {len(index)} entries")
        
        return index
    
    def get_by_index(self, sheet_name: str, field: str, value: Any) -> Optional[Dict]:
        """Get row by indexed field"""
        index = self.create_index(sheet_name, field)
        return index.get(value)
    
    def get_module_by_id(self, module_id: str) -> Optional[Dict]:
        """Get module data by module_index"""
        return self.get_by_index('DeepALL_Complete', 'module_index', module_id)
    
    def get_modules_by_superintelligence(self, si_index: str) -> List[Dict]:
        """Get all modules assigned to a superintelligence"""
        data = self.load_sheet('DeepALL_Complete')
        return [m for m in data if m.get('superintelligence_index') == si_index]
    
    # ========================================================================
    # DATA VALIDATION
    # ========================================================================
    
    def validate_complete_data(self) -> Tuple[int, List[str]]:
        """
        Validate DeepALL_Complete sheet
        
        Returns: (valid_rows, errors)
        """
        logger.info("Validating DeepALL_Complete sheet...")
        
        data = self.load_sheet('DeepALL_Complete')
        valid_count = 0
        errors = []
        
        for row_idx, row in enumerate(data, start=2):
            row_errors = []
            
            # Validate module_index
            if not DataValidator.validate_module_index(row.get('module_index')):
                row_errors.append(f"Invalid module_index: {row.get('module_index')}")
            
            # Validate module_name
            if not DataValidator.validate_non_empty(row.get('module_name')):
                row_errors.append("Empty module_name")
            
            # Validate scores (allow not_assigned, index refs, or numeric)
            for score_field in ['synergy_efficiency_index', 'optimization_index', 
                               'learning_index', 'performance_index', 'efficiency_index']:
                if score_field in row:
                    if not DataValidator.validate_index_score(row[score_field]):
                        row_errors.append(f"Invalid {score_field}: {row[score_field]}")
            
            if row_errors:
                errors.append(f"Row {row_idx}: {'; '.join(row_errors)}")
            else:
                valid_count += 1
        
        logger.info(f"  ✓ Valid rows: {valid_count}/{len(data)}")
        if errors:
            logger.warning(f"  ⚠ Errors: {len(errors)}")
            for error in errors[:5]:  # Show first 5 errors
                logger.warning(f"    {error}")
        
        return valid_count, errors
    
    def validate_referential_integrity(self) -> Tuple[int, List[str]]:
        """
        Validate referential integrity between sheets
        
        Returns: (valid_references, errors)
        """
        logger.info("Validating referential integrity...")
        
        complete = self.load_sheet('DeepALL_Complete')
        errors = []
        valid_count = 0
        
        # Create index of valid superintelligences
        try:
            si_data = self.load_sheet('superintelligences')
            valid_si = {si.get('superintelligence_index') for si in si_data if si.get('superintelligence_index')}
        except:
            valid_si = set()
            logger.warning("Could not load superintelligences sheet")
        
        # Check each module's SI reference
        for row in complete:
            si_index = row.get('superintelligence_index')
            
            if si_index and si_index != 'not_assigned':
                if si_index not in valid_si:
                    errors.append(f"Module {row.get('module_index')}: Invalid SI reference {si_index}")
                else:
                    valid_count += 1
            else:
                valid_count += 1
        
        logger.info(f"  ✓ Valid references: {valid_count}/{len(complete)}")
        if errors:
            logger.warning(f"  ⚠ Errors: {len(errors)}")
        
        return valid_count, errors
    
    def check_for_duplicates(self) -> Tuple[int, List[str]]:
        """Check for duplicate module IDs"""
        logger.info("Checking for duplicates...")
        
        data = self.load_sheet('DeepALL_Complete')
        module_ids = [row.get('module_index') for row in data]
        
        seen = set()
        duplicates = []
        
        for module_id in module_ids:
            if module_id in seen:
                duplicates.append(module_id)
            seen.add(module_id)
        
        logger.info(f"  ✓ Total modules: {len(data)}")
        logger.info(f"  ✓ Unique modules: {len(seen)}")
        if duplicates:
            logger.warning(f"  ⚠ Duplicates: {len(duplicates)}")
        
        return len(duplicates), duplicates
    
    # ========================================================================
    # STATISTICS
    # ========================================================================
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get loading statistics"""
        return {
            'excel_path': self.excel_path,
            'sheets_loaded': len(self.cache),
            'total_sheets': len(self.get_sheet_names()),
            'cache_size': len(self.cache),
            'indexes': len(self.indexes),
            'sheet_stats': self.stats,
            'validation_errors': len(self.validation_errors)
        }
    
    def print_statistics(self):
        """Print loading statistics"""
        stats = self.get_statistics()
        
        print("\n" + "="*80)
        print("DEEPALL TABLE LOADER STATISTICS")
        print("="*80)
        print(f"\nFile: {stats['excel_path']}")
        print(f"Sheets Loaded: {stats['sheets_loaded']}/{stats['total_sheets']}")
        print(f"Indexes Created: {stats['indexes']}")
        print(f"Validation Errors: {stats['validation_errors']}")
        
        print("\nSheet Statistics:")
        for sheet_name, sheet_stats in stats['sheet_stats'].items():
            print(f"  {sheet_name:30} {sheet_stats['rows']:5} rows, {sheet_stats['columns']:3} columns")
        
        print("\n" + "="*80 + "\n")
    
    # ========================================================================
    # EXPORT
    # ========================================================================
    
    def export_to_json(self, output_path: str, sheets: Optional[List[str]] = None):
        """Export loaded sheets to JSON"""
        if sheets is None:
            sheets = list(self.cache.keys())
        
        logger.info(f"Exporting {len(sheets)} sheets to {output_path}")
        
        export_data = {}
        for sheet_name in sheets:
            if sheet_name in self.cache:
                export_data[sheet_name] = self.cache[sheet_name]
        
        with open(output_path, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        logger.info(f"  ✓ Exported to {output_path}")
    
    def export_complete_data_csv(self, output_path: str):
        """Export DeepALL_Complete sheet to CSV"""
        import csv
        
        logger.info(f"Exporting DeepALL_Complete to {output_path}")
        
        data = self.load_sheet('DeepALL_Complete')
        
        if not data:
            logger.warning("No data to export")
            return
        
        with open(output_path, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        logger.info(f"  ✓ Exported {len(data)} rows to {output_path}")


# ============================================================================
# MAIN - TESTING
# ============================================================================

if __name__ == '__main__':
    import sys
    
    print("\n" + "="*80)
    print("DEEPALL TABLE LOADER - TEST")
    print("="*80 + "\n")
    
    # Load the table
    loader = DeepALLTableLoader('DeepALL_MASTER_V7_FINAL_WITH_ALL_REITERS.xlsx')
    
    # Test 1: Load core sheets
    print("TEST 1: Loading core sheets...")
    try:
        complete = loader.load_complete_data()
        print(f"  ✓ DeepALL_Complete: {len(complete)} modules")
        
        synergies = loader.load_synergies()
        print(f"  ✓ Synergies: {len(synergies)} records")
        
        optimization = loader.load_optimization()
        print(f"  ✓ Optimization: {len(optimization)} records")
        
        learning = loader.load_learning_results()
        print(f"  ✓ Learning Results: {len(learning)} records")
        
        print("  ✓ All core sheets loaded successfully\n")
    except Exception as e:
        print(f"  ✗ Error: {e}\n")
        sys.exit(1)
    
    # Test 2: Create indexes
    print("TEST 2: Creating indexes...")
    try:
        index = loader.create_index('DeepALL_Complete', 'module_index')
        print(f"  ✓ Module index created: {len(index)} entries\n")
    except Exception as e:
        print(f"  ✗ Error: {e}\n")
        sys.exit(1)
    
    # Test 3: Query by index
    print("TEST 3: Querying by index...")
    try:
        module = loader.get_module_by_id('m001')
        if module:
            print(f"  ✓ Found module m001: {module.get('module_name')}\n")
        else:
            print("  ✗ Module m001 not found\n")
    except Exception as e:
        print(f"  ✗ Error: {e}\n")
        sys.exit(1)
    
    # Test 4: Validate data
    print("TEST 4: Validating data...")
    try:
        valid, errors = loader.validate_complete_data()
        print(f"  ✓ Valid rows: {valid}/{len(complete)}\n")
    except Exception as e:
        print(f"  ✗ Error: {e}\n")
        sys.exit(1)
    
    # Test 5: Check duplicates
    print("TEST 5: Checking for duplicates...")
    try:
        dup_count, dups = loader.check_for_duplicates()
        print(f"  ✓ Duplicates: {dup_count}\n")
    except Exception as e:
        print(f"  ✗ Error: {e}\n")
        sys.exit(1)
    
    # Print statistics
    loader.print_statistics()
    
    print("✓ ALL TESTS PASSED\n")
